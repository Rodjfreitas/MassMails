def enviarEmail(usuario, senha_usuario, assunto, destinatario, com_copia, conteudo, caminho_anexo, nome_destinatario, servidor, anexar):
    import os
    import smtplib
    # from email.message import EmailMessage
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email.utils import COMMASPACE
    from email import encoders

    from email.mime.application import MIMEApplication

    # configurar email, senha
    EMAIL_ADRESS = usuario
    EMAIL_PASSWORD = senha_usuario
    POSSUI_ANEXO = anexar

    # Criando uma lista para imprimir um aquivo de retorno
    envio = {}

    try:
        # Startar o servidor
        host = servidor
        port = '587'
        server = smtplib.SMTP(host, port)
        server.ehlo()
        server.starttls()
        try:
            server.login(EMAIL_ADRESS, EMAIL_PASSWORD)
        except:
            print('Login ou Senha incorretos. revise-os.')
            return 'Erro'

        # Criar um email
        msg = MIMEMultipart()
        msg['Subject'] = assunto
        msg['From'] = usuario
        msg['To'] = destinatario
        msg['cc'] = com_copia
        msg.attach(MIMEText(f'{conteudo}', 'plain'))

        try:
            # Anexar arquivos
            pdf_folder = caminho_anexo
            pdf_files = [os.path.join(pdf_folder, f) for f in os.listdir(
                pdf_folder) if f.endswith('.pdf')]
            qtd_files = 0

            for file in pdf_files:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(open(file, 'rb').read())
                encoders.encode_base64(part)
                part.add_header('content-Disposition', 'attachment',
                                filename=os.path.basename(file))
                msg.attach(part)
                qtd_files += 1
        except (FileNotFoundError):
            pdf_folder = ''
            qtd_files = 0

        # Enviar email
        if qtd_files > 0 or POSSUI_ANEXO == 'nao':
            server.sendmail(msg['From'], msg['To'], msg.as_string())
            server.quit()

        # Informações para relatório de envio
        envio['nome'] = nome_destinatario
        envio['assunto'] = assunto
        envio['email'] = destinatario
        envio['status'] = 'Enviado'
        if qtd_files == 0:
            if POSSUI_ANEXO == 'nao':
                envio['status'] = 'Enviado'
                envio['obs'] = 'E-mail enviado com a opção sem anexo'
                print(f'{"Email Enviado: SIM":<30}  --  {assunto:<100}')
            else:
                envio['status'] = 'Não Enviado'
                envio['obs'] = 'Não foi enviado arquivo(s). Vefificar Diretório Informado'
                print(
                    f'{"Email Enviado: NÃO":<30}  --  {assunto:<100}')
        else:
            if qtd_files == 1:
                envio['obs'] = f'{qtd_files} arquivo enviado'
            else:
                envio['obs'] = f'{qtd_files} arquivos enviados'
            print(
                f'{"Email Enviado: SIM":<30}  --  {assunto:<100}')
        return envio
    except:
        print(
            f'{"Email Enviado: NÃO":<30}  --  {assunto:<100}')
        envio['nome'] = nome_destinatario
        envio['assunto'] = assunto
        if msg['To'] is None:
            envio['email'] = 'Inválido / Não informado'
        else:
            envio['email'] = destinatario
        envio['status'] = 'Não Enviado'
        if qtd_files != 0:
            envio['obs'] = 'Email Inválido ou não informado'
        else:
            envio['obs'] = 'Email Inválido ou não informado. Verificar também Diretório Informado.'
        return envio


def saudacao():
    import datetime
    horas = datetime.datetime.now().hour
    if 0 <= horas <= 11:
        return 'Bom dia'
    if horas <= 17:
        return 'Boa tarde'
    return 'Boa noite'


def buscarnoArquivo(arquivo):
    import openpyxl
    try:
        # Abre o arquivo excel
        workbook = openpyxl.load_workbook(arquivo)

        # Seleciona a planilha desejada
        sheet = workbook['Planilha1']

        # Itera sobre todas as células da planilha
        valores = []
        linhas = {}
        i = 0
        for row in sheet.iter_rows():
            for cell in row:
                linhas[i] = cell.value
                i += 1
            valores.append(linhas.copy())
            linhas.clear()
            i = 0
    except:
        return 0
    return valores


def salvanoArquivo(lista):
    import os
    import openpyxl

    nome_do_arquivo = 'Relatorio_emails_disparados.xlsx'
    arquivo = openpyxl.Workbook()
    planilha = arquivo.active
    planilha.title = 'Relatório de emails disparados'
    cabecalho_colunas = ['Nome', 'Assunto', 'Email', 'Status', 'Obs:']

    for coluna, cabecalho in enumerate(cabecalho_colunas, start=1):
        planilha.cell(row=1, column=coluna).value = cabecalho

    for linha, dados_linha in enumerate(lista, start=2):
        for coluna, dado in enumerate(dados_linha.values(), start=1):
            planilha.cell(row=linha, column=coluna).value = dado

    # verifica se já existe um arquivo com mesmo nome.
    if os.path.isfile(nome_do_arquivo):
        # se o arquivo existe, exclui-o
        os.remove(nome_do_arquivo)
    arquivo.save(nome_do_arquivo)
    print('Foi Salvo um arquivo com a relação de envios na pasta do aplicativo Mass&Mails.')


def imprimirLista(lista):
    for pos, valor in enumerate(lista):
        try:
            if pos == 0:
                print(
                    f'{valor[0]:^46}{valor[1]:^60}')
                continue
            print(
                f'{pos:<3} - {valor[0]:<40}{valor[1]:^60}')
        except:
            print(
                f'{pos:<3} - IMPORTANTE: Revise os dados de {valor[0]} na planilha origem.\n')
            continue
